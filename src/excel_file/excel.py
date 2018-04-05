#!/usr/bin/env python
# coding:utf-8

from openpyxl import load_workbook
from openpyxl import Workbook
import json


class OperateExcel:

    def __init__(self, load_filename, store_filename):
        self.load_filename = load_filename
        self.store_filename = store_filename

    # 导入excel文件中的数据
    def load_excel(self):
        # 打开一个workbook
        wb = load_workbook(self.load_filename)

        print(wb.sheetnames)
        sheetnames = wb.sheetnames
        first_sheet = wb[sheetnames[0]]
        print first_sheet

        print "Work Sheet Titile:", first_sheet.title
        print "Work Sheet Max Rows:", first_sheet.max_row
        print "Work Sheet Min Rows:", first_sheet.min_row
        print "Work Sheet Max Cols:", first_sheet.max_column
        print "Work Sheet Min Cols:", first_sheet.min_column

        # 建立存储数据的字典
        data_store = {}

        # 将数据存到字典中
        for row in range(first_sheet.min_row, first_sheet.max_row + 1):
            temp_list = []
            for column in range(first_sheet.min_column, first_sheet.max_column + 1):
                content = first_sheet.cell(row=row, column=column).value
                temp_list.append(content)
            data_store[row] = temp_list

        # 打印字典数据个数
        print 'Total:%d' % len(data_store)
        content_json = json.dumps(data_store, encoding="UTF-8", ensure_ascii=False)
        print content_json
        return content_json

    # 将数据导出到excel文件中
    def export_excel(self, data=None):
        print("export_excel")
        wb = Workbook()
        first_sheet = wb.active
        first_sheet.title = u"哈哈"
        if data is not None:
            content = json.loads(data, encoding="UTF-8")
            print content

            row_list = {}
            for row in content.iterkeys():
                row_list[int(row)] = row

            print row_list
            for row in range(1, len(row_list) + 1):
                print u'second_row : %s' % row
                print u'second_content[row_list[row]] : %s' % content[row_list[row]]
                first_sheet.append(content[row_list[row]])

        wb.save(self.store_filename)
        print "保存成功"
        return
